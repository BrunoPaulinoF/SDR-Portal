#!/usr/bin/env python3
"""Monta a planilha final a partir do JSON ranqueado do rank-leads-gastro.mjs.

    python3 scripts/leads-gastro-xlsx.py [entrada.json] [saida.xlsx]

Tres abas: Leads (a lista para trabalhar), Resumo (contagens por COUNTIFS, que
recalculam sozinhas se voce apagar linhas) e Criterio (como o score foi montado).
"""
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ENTRADA = sys.argv[1] if len(sys.argv) > 1 else "local-secrets/leads-ranked.json"
SAIDA = sys.argv[2] if len(sys.argv) > 2 else "docs/leads/leads-gastro-delivery-sp.xlsx"

COLUNAS = [
    ("rank", "#", 5), ("score", "Score", 7), ("nome", "Nome da empresa", 34),
    ("tipo", "Tipo", 26), ("categoria", "Segmento", 20),
    ("cidade", "Cidade", 17), ("estado", "Estado", 8), ("bairro", "Bairro", 18),
    ("endereco", "Endereco", 40),
    ("telefone", "Telefone (formatado)", 20), ("telefone_e164", "WhatsApp", 16),
    ("whatsapp_provavel", "Tem WhatsApp?", 15),
    ("delivery", "Delivery", 22), ("plataformas", "Plataformas", 20),
    ("nota", "Nota", 7), ("avaliacoes", "Avaliacoes", 11),
    ("site", "Site", 30), ("maps", "Google Maps", 22),
    ("porque", "Por que entrou", 60), ("ressalvas", "Ressalvas", 50),
]

# Estes cabecalhos sao lidos pelo import de leads do portal, que casa por alias
# (src/modules/leads/lead-importer.ts): "Nome da empresa", "WhatsApp", "Segmento",
# "Cidade" e "Estado". Renomear qualquer um deles quebra o import automatico.
# "Telefone (formatado)" e "Tem WhatsApp?" sao propositalmente NAO-aliases, para
# nao disputarem o mapeamento com a coluna do numero.

IDX = {chave: i for i, (chave, _, _) in enumerate(COLUNAS, start=1)}


def col(chave):
    """Letra da coluna a partir da chave, para as formulas nunca sairem do lugar."""
    return get_column_letter(IDX[chave])

AZUL = "1F3864"
FAIXA = "F2F5FA"
VERDE = "C6EFCE"
AMARELO = "FFEB9C"
VERMELHO = "FFC7CE"

FONTE = "Arial"
borda = Border(bottom=Side(style="thin", color="D0D7E5"))


def estilo_base(ws, n_cols, n_rows):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{n_rows}"
    for c, (_, titulo, larg) in enumerate(COLUNAS, start=1):
        cel = ws.cell(row=1, column=c, value=titulo)
        cel.font = Font(name=FONTE, bold=True, color="FFFFFF", size=10)
        cel.fill = PatternFill("solid", fgColor=AZUL)
        cel.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = larg
    ws.row_dimensions[1].height = 30


def aba_leads(wb, leads):
    ws = wb.create_sheet("Leads")
    estilo_base(ws, len(COLUNAS), len(leads) + 1)
    for r, lead in enumerate(leads, start=2):
        for c, (chave, _, _) in enumerate(COLUNAS, start=1):
            valor = lead.get(chave, "")
            cel = ws.cell(row=r, column=c, value=valor if valor != "" else None)
            cel.font = Font(name=FONTE, size=9)
            cel.border = borda
            cel.alignment = Alignment(vertical="top", wrap_text=chave in ("porque", "ressalvas", "endereco"))
            if r % 2 == 0:
                cel.fill = PatternFill("solid", fgColor=FAIXA)
        # Score colorido: verde = trabalhar primeiro, vermelho = so se sobrar tempo.
        s = lead.get("score", 0)
        cor = VERDE if s >= 75 else AMARELO if s >= 55 else VERMELHO
        ws.cell(row=r, column=IDX["score"]).fill = PatternFill("solid", fgColor=cor)
        ws.cell(row=r, column=IDX["score"]).font = Font(name=FONTE, size=9, bold=True)
        # WhatsApp duvidoso e o que mais queima base: destaca.
        if lead.get("whatsapp_provavel") != "Sim":
            ws.cell(row=r, column=IDX["whatsapp_provavel"]).fill = PatternFill("solid", fgColor=VERMELHO)
        for coluna, chave in ((IDX["site"], "site"), (IDX["maps"], "maps")):
            url = lead.get(chave)
            if url:
                cel = ws.cell(row=r, column=coluna)
                cel.hyperlink = url
                cel.value = "abrir" if chave == "maps" else url[:60]
                cel.font = Font(name=FONTE, size=9, color="0563C1", underline="single")
    return ws


def aba_resumo(wb, leads):
    ws = wb.create_sheet("Resumo", 0)
    n = len(leads) + 1
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12

    def titulo(row, texto):
        c = ws.cell(row=row, column=1, value=texto)
        c.font = Font(name=FONTE, bold=True, size=12, color=AZUL)

    def linha(row, rotulo, formula, pct=True):
        ws.cell(row=row, column=1, value=rotulo).font = Font(name=FONTE, size=10)
        c = ws.cell(row=row, column=2, value=formula)
        c.font = Font(name=FONTE, size=10)
        if pct:
            p = ws.cell(row=row, column=3, value=f"=IFERROR(B{row}/$B$3,0)")
            p.number_format = "0.0%"
            p.font = Font(name=FONTE, size=10, color="595959")

    titulo(1, "Leads gastronomia + delivery — Sao Paulo, ABC e interior")
    ws.cell(row=2, column=1, value="Fonte: Google Maps via Apify (compass/crawler-google-places)").font = Font(
        name=FONTE, size=9, italic=True, color="595959")

    linha(3, "Total de leads na planilha", f'=COUNTA(Leads!{col("rank")}2:{col("rank")}{n})', pct=False)
    linha(4, "Com celular (WhatsApp provavel)", f'=COUNTIF(Leads!{col("whatsapp_provavel")}2:{col("whatsapp_provavel")}{n},"Sim")')
    linha(5, "Delivery confirmado por plataforma de pedido", f'=COUNTIF(Leads!{col("delivery")}2:{col("delivery")}{n},"Confirmado*")')
    linha(6, "Entrega marcada no perfil do Google", f'=COUNTIF(Leads!{col("delivery")}2:{col("delivery")}{n},"Sim*")')
    linha(7, "Rede ou franquia", f'=COUNTIF(Leads!{col("tipo")}2:{col("tipo")}{n},"Rede/franquia*")+COUNTIF(Leads!{col("tipo")}2:{col("tipo")}{n},"Franquia*")')
    linha(8, "Score alto (>=75): comecar por aqui", f'=COUNTIF(Leads!{col("score")}2:{col("score")}{n},">=75")')
    linha(9, "Prontos para abordagem (celular + plataforma de pedido)",
          f'=COUNTIFS(Leads!{col("whatsapp_provavel")}2:{col("whatsapp_provavel")}{n},"Sim",Leads!{col("delivery")}2:{col("delivery")}{n},"Confirmado*")')

    titulo(11, "Por cidade")
    cidades = sorted({l.get("cidade", "") for l in leads if l.get("cidade")})
    for i, cidade in enumerate(cidades):
        r = 12 + i
        ws.cell(row=r, column=1, value=cidade).font = Font(name=FONTE, size=10)
        c = ws.cell(row=r, column=2, value=f'=COUNTIF(Leads!{col("cidade")}2:{col("cidade")}{n},A{r})')
        c.font = Font(name=FONTE, size=10)
        p = ws.cell(row=r, column=3, value=f"=IFERROR(B{r}/$B$3,0)")
        p.number_format = "0.0%"
        p.font = Font(name=FONTE, size=10, color="595959")

    r = 13 + len(cidades)
    titulo(r, "Como usar")
    for j, t in enumerate([
        "Trabalhe de cima para baixo: a aba Leads ja vem ordenada por score.",
        "A coluna E.164 e a que entra direto no import de leads do portal.",
        "WhatsApp? em vermelho = numero fixo. Confira antes de gastar disparo.",
        "Ressalvas explica o que pesou contra cada lead.",
        "As contagens acima sao formulas: apague linhas e elas se ajustam.",
    ]):
        ws.cell(row=r + 1 + j, column=1, value=f"• {t}").font = Font(name=FONTE, size=10)


def aba_criterio(wb):
    ws = wb.create_sheet("Criterio")
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 78
    cab = ["Fator", "Pontos", "Por que"]
    for c, t in enumerate(cab, start=1):
        cel = ws.cell(row=1, column=c, value=t)
        cel.font = Font(name=FONTE, bold=True, color="FFFFFF", size=10)
        cel.fill = PatternFill("solid", fgColor=AZUL)
    regras = [
        ("Telefone e celular", "+30", "18 dos 82 leads da Mariana morreram como invalid_phone. Numero fixo nao tem WhatsApp."),
        ("Telefone fixo", "+4", "Fica na lista, mas so vale disparo depois de conferir."),
        ("Delivery confirmado (plataforma)", "+25", "Site aponta para iFood, Anota AI, Cardapio Web, Goomer, Menudino ou Saipos: operacao digital rodando."),
        ("Entrega no perfil do Google", "+16", "81% dos lugares marcam entrega, entao isso qualifica mas quase nao diferencia."),
        ("Delivery provavel", "+10", "So o nome ou a categoria sugerem entrega."),
        ("80 a 900 avaliacoes", "+20", "Faixa de porte da oferta: pequeno e medio que ja vende."),
        ("25 a 79 avaliacoes", "+14", "Casa menor, ainda dentro do publico."),
        ("900 a 2500 avaliacoes", "+9", "Operacao grande: costuma ja ter stack propria."),
        ("Acima de 2500 avaliacoes", "+3", "Grande demais para a oferta."),
        ("Menos de 25 avaliacoes", "+3", "Casa nova ou pouco ativa: risco de nao ter volume."),
        ("Nota entre 4,0 e 4,9", "+10", "Operacao cuidada e com gestao presente."),
        ("Rede/franquia na base", "+15", "Duas ou mais unidades da mesma marca apareceram na varredura."),
        ("Franquia de marca conhecida", "+15", "Rede de porte medio onde o franqueado decide a propria operacao."),
        ("Site proprio", "+5", "Sinal de que a casa ja investe em canal digital."),
        ("Rede nacional", "-25", "O numero publicado e central ou robo de pedido, e a unidade nao decide sozinha."),
    ]
    for r, (f, p, m) in enumerate(regras, start=2):
        for c, v in enumerate((f, p, m), start=1):
            cel = ws.cell(row=r, column=c, value=v)
            cel.font = Font(name=FONTE, size=9)
            cel.alignment = Alignment(vertical="top", wrap_text=(c == 3))
            if r % 2 == 0:
                cel.fill = PatternFill("solid", fgColor=FAIXA)

    r = len(regras) + 3
    ws.cell(row=r, column=1, value="Descartados antes do ranking").font = Font(name=FONTE, bold=True, size=11, color=AZUL)
    for j, t in enumerate([
        "Lugares fechados em definitivo ou temporariamente.",
        "Sem telefone publicado (nao ha como abordar).",
        "Telefone repetido (a mesma casa aparecendo em varias buscas).",
        "Categoria fora do ramo (hotel, posto, mercado, farmacia).",
    ]):
        ws.cell(row=r + 1 + j, column=1, value=f"• {t}").font = Font(name=FONTE, size=9)


def main():
    with open(ENTRADA, encoding="utf-8") as fh:
        leads = json.load(fh)
    wb = Workbook()
    wb.remove(wb.active)
    aba_leads(wb, leads)
    aba_resumo(wb, leads)
    aba_criterio(wb)
    wb.move_sheet("Leads", offset=-1)
    import os
    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    wb.save(SAIDA)
    print(f"{len(leads)} leads -> {SAIDA}")


main()
